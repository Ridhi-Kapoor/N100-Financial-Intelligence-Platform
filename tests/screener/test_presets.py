"""
Unit and integration tests for the screener presets module.
"""

from pathlib import Path
import pandas as pd
import pytest

from src.screener.presets import (
    apply_preset,
    apply_quality_compounder_filter,
    apply_value_pick_filter,
    apply_growth_accelerator_filter,
    apply_dividend_champion_filter,
    apply_debt_free_blue_chip_filter,
    apply_turnaround_watch_filter,
    screen_quality_compounders,
    screen_value_picks,
    screen_growth_accelerators,
    screen_dividend_champions,
    screen_debt_free_blue_chips,
    screen_turnaround_watch,
    load_and_score_data
)

PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
DB_FILE = PROJECT_ROOT / "data" / "db" / "nifty100.db"


def test_preset_invalid_name():
    """Test that applying an invalid preset name raises ValueError."""
    with pytest.raises(ValueError):
        apply_preset("Non-Existent Preset", DB_FILE, year=2024)


def test_quality_compounder_mock():
    """Test the Quality Compounder preset filters on mock data."""
    # Columns required by quality compounder:
    # ROE, Debt-to-Equity, Free Cash Flow, Operating Profit Margin, Interest Coverage Ratio, composite_quality_score
    data = {
        "company_id": ["A", "B", "C"],
        "company_name": ["Comp A", "Comp B", "Comp C"],
        "year": ["2024", "2024", "2024"],
        "return_on_equity_pct": [20.0, 10.0, 18.0],
        "debt_to_equity": [0.2, 0.1, 0.9],  # C fails Debt-to-Equity (> 0.5)
        "free_cash_flow_cr": [100.0, 50.0, 20.0],
        "operating_profit_margin_pct": [15.0, 8.0, 20.0],  # B fails OPM (< 12%)
        "interest_coverage": [5.0, 4.0, 6.0],
        "composite_quality_score": [1.5, 1.2, 1.8],
        "broad_sector": ["Tech", "Tech", "Tech"]
    }
    df = pd.DataFrame(data)
    res = apply_quality_compounder_filter(df)
    
    # Only A should pass
    assert len(res) == 1
    assert res.iloc[0]["company_id"] == "A"


def test_value_pick_mock():
    """Test the Value Pick preset filters on mock data."""
    # PE <= 35, PB <= 5, ROE >= 12, D/E <= 1.2
    data = {
        "company_id": ["A", "B", "C"],
        "company_name": ["Comp A", "Comp B", "Comp C"],
        "year": ["2024", "2024", "2024"],
        "pe_ratio": [20.0, 40.0, 15.0],  # B fails PE (> 35)
        "pb_ratio": [3.0, 2.0, 6.0],   # C fails PB (> 5)
        "return_on_equity_pct": [15.0, 18.0, 14.0],
        "debt_to_equity": [0.5, 0.4, 0.3],
        "composite_quality_score": [1.0, 1.2, 1.1],
        "broad_sector": ["Tech", "Tech", "Tech"]
    }
    df = pd.DataFrame(data)
    res = apply_value_pick_filter(df)
    
    # Only A should pass
    assert len(res) == 1
    assert res.iloc[0]["company_id"] == "A"


def test_growth_accelerator_mock():
    """Test the Growth Accelerator preset filters on mock data."""
    # Revenue CAGR 5Y >= 12, PAT CAGR 5Y >= 12, ROE >= 12, OPM >= 10
    data = {
        "company_id": ["A", "B"],
        "company_name": ["Comp A", "Comp B"],
        "year": ["2024", "2024"],
        "revenue_cagr_5yr": [15.0, 8.0],  # B fails revenue CAGR
        "pat_cagr_5yr": [20.0, 14.0],
        "return_on_equity_pct": [18.0, 15.0],
        "operating_profit_margin_pct": [14.0, 12.0],
        "composite_quality_score": [1.0, 1.2],
        "broad_sector": ["Tech", "Tech"]
    }
    df = pd.DataFrame(data)
    res = apply_growth_accelerator_filter(df)
    
    assert len(res) == 1
    assert res.iloc[0]["company_id"] == "A"


def test_dividend_champion_mock():
    """Test the Dividend Champion preset filters on mock data."""
    # Dividend Yield >= 2.0, ROE >= 10, D/E <= 1.2
    data = {
        "company_id": ["A", "B"],
        "company_name": ["Comp A", "Comp B"],
        "year": ["2024", "2024"],
        "dividend_yield_pct": [2.5, 1.2],  # B fails Dividend Yield
        "return_on_equity_pct": [12.0, 15.0],
        "debt_to_equity": [0.6, 0.4],
        "composite_quality_score": [1.0, 1.2],
        "broad_sector": ["Tech", "Tech"]
    }
    df = pd.DataFrame(data)
    res = apply_dividend_champion_filter(df)
    
    assert len(res) == 1
    assert res.iloc[0]["company_id"] == "A"


def test_debt_free_blue_chip_mock():
    """Test the Debt-Free Blue Chip preset filters on mock data."""
    # Market Cap >= 20000, ROE >= 12, (is_debt_free is True OR D/E <= 0.05)
    data = {
        "company_id": ["A", "B", "C"],
        "company_name": ["Comp A", "Comp B", "Comp C"],
        "year": ["2024", "2024", "2024"],
        "market_cap_crore": [25000.0, 15000.0, 30000.0],  # B fails Market Cap (< 20000)
        "return_on_equity_pct": [15.0, 14.0, 16.0],
        "debt_to_equity": [0.03, 0.01, 0.20],            # C fails Debt-to-Equity (> 0.05) and interest check
        "interest": [2.0, 0.0, 10.0],
        "composite_quality_score": [1.0, 1.2, 1.1],
        "broad_sector": ["Tech", "Tech", "Tech"]
    }
    df = pd.DataFrame(data)
    res = apply_debt_free_blue_chip_filter(df)
    
    # Only A should pass (B failed Market Cap, C failed leverage constraints)
    assert len(res) == 1
    assert res.iloc[0]["company_id"] == "A"


def test_turnaround_watch_mock():
    """Test the Turnaround Watch preset filters on mock data."""
    # D/E (t) < D/E (t-1), ROE >= 10, Net Profit >= 100
    df_all = pd.DataFrame({
        "company_id": ["A", "A", "B", "B"],
        "year": ["2023", "2024", "2023", "2024"],
        "debt_to_equity": [0.8, 0.5, 0.3, 0.4],  # A declining (0.8 -> 0.5), B increasing (0.3 -> 0.4)
        "return_on_equity_pct": [12.0, 14.0, 10.0, 15.0],
        "net_profit": [150.0, 200.0, 120.0, 180.0],
        "composite_quality_score": [1.0, 1.1, 1.2, 1.3],
        "broad_sector": ["Tech", "Tech", "Tech", "Tech"]
    })
    
    df_2024 = df_all[df_all["year"] == "2024"].copy()
    res = apply_turnaround_watch_filter(df_2024, df_all)
    
    # Only A should pass because its D/E declined YoY
    assert len(res) == 1
    assert res.iloc[0]["company_id"] == "A"


def test_presets_integration_with_db():
    """Integration test checking that all presets return valid outputs on real DB data."""
    if not DB_FILE.exists():
        pytest.skip("SQLite database nifty100.db not found. Skipping integration test.")
        
    presets = [
        "Quality Compounder", "Value Pick", "Growth Accelerator",
        "Dividend Champion", "Debt-Free Blue Chip", "Turnaround Watch"
    ]
    
    for preset_name in presets:
        df_res = apply_preset(preset_name, DB_FILE, year=2024)
        
        # Verify output structure
        assert isinstance(df_res, pd.DataFrame)
        assert not df_res.empty
        assert "company_id" in df_res.columns
        assert "company_name" in df_res.columns
        
        # Verify counts are within required limits (5 to 50)
        assert 5 <= len(df_res) <= 50, f"Preset '{preset_name}' yielded {len(df_res)} companies, which is outside [5, 50] range."


def test_composite_score_calculation_integration():
    """Integration test checking that composite scores are calculated in [0, 100]."""
    if not DB_FILE.exists():
        pytest.skip("SQLite database nifty100.db not found. Skipping integration test.")
        
    df_scored = load_and_score_data(DB_FILE, year=2024)
    assert not df_scored.empty
    assert "composite_quality_score" in df_scored.columns
    
    # Check that scores are numeric and within 0-100 bounds
    scores = df_scored["composite_quality_score"]
    assert scores.notna().all()
    assert scores.min() >= 0.0
    assert scores.max() <= 100.0


def test_excel_generation_integration(tmp_path):
    """Integration test checking that the Excel spreadsheet is generated successfully."""
    if not DB_FILE.exists():
        pytest.skip("SQLite database nifty100.db not found. Skipping integration test.")
        
    output_excel = tmp_path / "screener_output.xlsx"
    from src.screener.excel_export import generate_screener_excel
    
    generate_screener_excel(DB_FILE, output_excel, year=2024)
    assert output_excel.exists()
    
    import openpyxl
    wb = openpyxl.load_workbook(output_excel)
    expected_sheets = [
        "Quality Compounder",
        "Value Pick",
        "Growth Accelerator",
        "Dividend Champion",
        "Debt-Free Blue Chip",
        "Turnaround Watch"
    ]
    for sheet in expected_sheets:
        assert sheet in wb.sheetnames
        ws = wb[sheet]
        assert ws.max_column == 20
        # Header row + at least 5 companies
        assert ws.max_row >= 6

