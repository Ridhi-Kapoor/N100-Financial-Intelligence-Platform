"""
Unit tests for src/analytics/peer_report.py module.
"""

from pathlib import Path
import openpyxl

from src.analytics.peer_report import (
    load_peer_comparison_dataset,
    generate_peer_comparison_excel,
    KPI_MAPPING,
)

PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
DB_PATH = PROJECT_ROOT / "data" / "db" / "nifty100.db"
OUTPUT_FILE = PROJECT_ROOT / "output" / "peer_comparison.xlsx"


def test_load_peer_comparison_dataset():
    """Verify loading dataset for peer comparison report."""
    df = load_peer_comparison_dataset(db_path=DB_PATH)
    assert not df.empty
    assert "peer_group_name" in df.columns
    assert "is_benchmark" in df.columns
    assert df["peer_group_name"].nunique() == 11

    for col in KPI_MAPPING:
        assert col in df.columns


def test_generate_peer_comparison_excel():
    """Verify Excel report generation and formatting."""
    out_path = generate_peer_comparison_excel(db_path=DB_PATH, output_path=OUTPUT_FILE)
    assert out_path.exists()
    assert out_path.stat().st_size > 0

    wb = openpyxl.load_workbook(out_path)
    assert len(wb.sheetnames) == 11

    # Check a specific sheet
    ws = wb["Private Banks"]
    assert ws.max_column == 42  # 2 ID cols + 20 KPIs + 20 Pct Ranks

    # Check benchmark row fill for HDFCBANK (row 2)
    benchmark_cell = ws.cell(row=2, column=1)
    assert benchmark_cell.value == "HDFCBANK"
    assert benchmark_cell.fill.start_color.rgb == "00FFE699"

    # Check summary row at bottom
    summary_id = ws.cell(row=ws.max_row, column=1).value
    summary_name = ws.cell(row=ws.max_row, column=2).value
    assert summary_id == "PEER MEDIAN"
    assert summary_name == "Peer Group Median"
    assert ws.cell(row=ws.max_row, column=1).fill.start_color.rgb == "00D9E1F2"
