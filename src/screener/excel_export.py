"""
Excel Export Module for Stock Screener.

This module handles exporting the filtered stocks for each preset to a 
multi-sheet Excel workbook (`output/screener_output.xlsx`). It applies 
professional styling and conditional formatting (green/red cell highlights) 
based on the thresholds of each preset.
"""

from pathlib import Path
from typing import Dict, Any, Union, List
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.screener.presets import (
    load_and_score_data,
    screen_quality_compounders,
    screen_value_picks,
    screen_growth_accelerators,
    screen_dividend_champions,
    screen_debt_free_blue_chips,
    screen_turnaround_watch
)
from src.screener.engine import is_debt_free, is_financial_sector

# Header fill and style
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")  # Dark blue
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

# Conditional formatting styles (Soft Green and Soft Red)
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
GREEN_FONT = Font(name="Calibri", size=11, color="006100", bold=True)

RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
RED_FONT = Font(name="Calibri", size=11, color="9C0006", bold=True)

# Normal text font
REGULAR_FONT = Font(name="Calibri", size=11)

# Border styles
THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)

# Standard list of 20 KPI columns to output in the Excel sheets
KPI_COLUMNS = [
    "company_id",
    "company_name",
    "broad_sector",
    "market_cap_crore",
    "sales",
    "net_profit",
    "return_on_equity_pct",
    "roce_percentage",
    "net_profit_margin_pct",
    "operating_profit_margin_pct",
    "revenue_cagr_5yr",
    "pat_cagr_5yr",
    "free_cash_flow_cr",
    "cash_from_operations_cr",
    "fcf_cagr_5yr",
    "cfo_pat_ratio",
    "debt_to_equity",
    "interest_coverage",
    "dividend_yield_pct",
    "composite_quality_score"
]

# Rename mapping for user presentation in Excel
COLUMN_RENAME_MAP = {
    "company_id": "Ticker",
    "company_name": "Company Name",
    "broad_sector": "Sector",
    "market_cap_crore": "Market Cap (Cr)",
    "sales": "Revenue (Cr)",
    "net_profit": "Net Profit (Cr)",
    "return_on_equity_pct": "ROE %",
    "roce_percentage": "ROCE %",
    "net_profit_margin_pct": "NPM %",
    "operating_profit_margin_pct": "OPM %",
    "revenue_cagr_5yr": "Revenue CAGR 5Y %",
    "pat_cagr_5yr": "PAT CAGR 5Y %",
    "free_cash_flow_cr": "FCF (Cr)",
    "cash_from_operations_cr": "CFO (Cr)",
    "fcf_cagr_5yr": "FCF CAGR 5Y %",
    "cfo_pat_ratio": "CFO/PAT Ratio",
    "debt_to_equity": "Debt-to-Equity",
    "interest_coverage": "Interest Coverage",
    "dividend_yield_pct": "Dividend Yield %",
    "composite_quality_score": "Composite Score"
}


def apply_cell_style(cell, fill=None, font=REGULAR_FONT, alignment=None, border=THIN_BORDER, num_format=None):
    """Utility to safely apply multiple styles to an openpyxl cell."""
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    if num_format:
        cell.number_format = num_format


def get_column_rename_header(col_name: str) -> str:
    """Get presentation name of a column."""
    return COLUMN_RENAME_MAP.get(col_name, col_name)


def style_preset_sheet(ws, df: pd.DataFrame, preset_name: str):
    """
    Apply professional layout, alignments, borders, number formatting, 
    and conditional highlights to the cells of a worksheet.
    """
    # 1. Format the headers
    ws.row_dimensions[1].height = 26
    for col_idx in range(1, len(df.columns) + 1):
        cell = ws.cell(row=1, column=col_idx)
        apply_cell_style(
            cell, 
            fill=HEADER_FILL, 
            font=HEADER_FONT, 
            alignment=Alignment(horizontal="center", vertical="center", wrap_text=True)
        )
        
    # Columns positioning lookup
    col_map = {col_name: idx + 1 for idx, col_name in enumerate(df.columns)}
    
    # 2. Iterate through data rows
    for r_idx, row in enumerate(df.itertuples(index=False), start=2):
        ws.row_dimensions[r_idx].height = 20
        row_dict = row._asdict() if hasattr(row, "_asdict") else row._fields
        
        # We can construct a dict mapping col name to value
        row_vals = {col_name: getattr(row, col_name) for col_name in df.columns}
        
        is_fin = is_financial_sector(row_vals)
        is_df = is_debt_free(row_vals)
        
        for col_name, val in row_vals.items():
            col_idx = col_map[col_name]
            cell = ws.cell(row=r_idx, column=col_idx)
            
            # Default formatting & alignments
            align = Alignment(vertical="center")
            num_fmt = None
            
            # Numeric alignment & number formats
            if col_name in ["company_id", "broad_sector"]:
                align = Alignment(horizontal="center", vertical="center")
            elif col_name == "company_name":
                align = Alignment(horizontal="left", vertical="center")
            elif col_name in ["market_cap_crore", "sales", "net_profit", "free_cash_flow_cr", "cash_from_operations_cr"]:
                align = Alignment(horizontal="right", vertical="center")
                num_fmt = "#,##0.00"
            elif col_name in ["return_on_equity_pct", "roce_percentage", "net_profit_margin_pct", 
                              "operating_profit_margin_pct", "revenue_cagr_5yr", "pat_cagr_5yr", 
                              "fcf_cagr_5yr", "dividend_yield_pct", "composite_quality_score"]:
                align = Alignment(horizontal="right", vertical="center")
                num_fmt = "0.00"
            elif col_name in ["debt_to_equity", "cfo_pat_ratio"]:
                align = Alignment(horizontal="right", vertical="center")
                num_fmt = "0.00"
            elif col_name == "interest_coverage":
                if is_df or val is None:
                    align = Alignment(horizontal="center", vertical="center")
                else:
                    align = Alignment(horizontal="right", vertical="center")
                    num_fmt = "0.00"
                    
            apply_cell_style(cell, alignment=align, num_format=num_fmt)
            
            # Format "Debt Free" text in interest coverage
            if col_name == "interest_coverage" and is_df:
                cell.value = "Debt Free"
                
            # Apply conditional formatting highlights row-by-row
            apply_preset_conditional_formatting(cell, col_name, val, is_fin, is_df, preset_name, row_vals)
            
    # 3. Auto-fit column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if cell.row == 1:
                # Add extra padding for wrapped header text
                max_len = max(max_len, len(val_str) // 2 + 5)
            else:
                max_len = max(max_len, len(val_str) + 3)
        ws.column_dimensions[col_letter].width = max(max_len, 12)


def apply_preset_conditional_formatting(cell, col_name: str, val, is_fin: bool, is_df: bool, preset_name: str, row_vals: Dict[str, Any]):
    """Apply green/red colors to cells that have filters in the active preset."""
    if val is None or pd.isna(val):
        return
        
    try:
        f_val = float(val)
    except (ValueError, TypeError):
        f_val = None
        
    # --- Preset 1: Quality Compounder ---
    if preset_name == "Quality Compounder":
        if col_name == "return_on_equity_pct":
            if f_val is not None:
                apply_cell_style(cell, fill=GREEN_FILL if f_val >= 15.0 else RED_FILL, font=GREEN_FONT if f_val >= 15.0 else RED_FONT)
        elif col_name == "debt_to_equity":
            if is_fin:
                # Financials are exempt, show neutral green since they passed the filter
                apply_cell_style(cell, fill=GREEN_FILL, font=GREEN_FONT)
            elif f_val is not None:
                apply_cell_style(cell, fill=GREEN_FILL if f_val <= 0.5 else RED_FILL, font=GREEN_FONT if f_val <= 0.5 else RED_FONT)
        elif col_name == "free_cash_flow_cr":
            if f_val is not None:
                apply_cell_style(cell, fill=GREEN_FILL if f_val >= 0.0 else RED_FILL, font=GREEN_FONT if f_val >= 0.0 else RED_FONT)
        elif col_name == "operating_profit_margin_pct":
            if f_val is not None:
                apply_cell_style(cell, fill=GREEN_FILL if f_val >= 12.0 else RED_FILL, font=GREEN_FONT if f_val >= 12.0 else RED_FONT)
        elif col_name == "interest_coverage":
            if is_df:
                apply_cell_style(cell, fill=GREEN_FILL, font=GREEN_FONT)
            elif f_val is not None:
                apply_cell_style(cell, fill=GREEN_FILL if f_val >= 3.0 else RED_FILL, font=GREEN_FONT if f_val >= 3.0 else RED_FONT)
                
    # --- Preset 2: Value Pick ---
    elif preset_name == "Value Pick":
        if col_name == "pe_ratio":
            if f_val is not None:
                apply_cell_style(cell, fill=GREEN_FILL if f_val <= 35.0 else RED_FILL, font=GREEN_FONT if f_val <= 35.0 else RED_FONT)
        elif col_name == "pb_ratio":
            if f_val is not None:
                apply_cell_style(cell, fill=GREEN_FILL if f_val <= 5.0 else RED_FILL, font=GREEN_FONT if f_val <= 5.0 else RED_FONT)
        elif col_name == "return_on_equity_pct":
            if f_val is not None:
                apply_cell_style(cell, fill=GREEN_FILL if f_val >= 12.0 else RED_FILL, font=GREEN_FONT if f_val >= 12.0 else RED_FONT)
        elif col_name == "debt_to_equity":
            if is_fin:
                apply_cell_style(cell, fill=GREEN_FILL, font=GREEN_FONT)
            elif f_val is not None:
                apply_cell_style(cell, fill=GREEN_FILL if f_val <= 1.2 else RED_FILL, font=GREEN_FONT if f_val <= 1.2 else RED_FONT)
                
    # --- Preset 3: Growth Accelerator ---
    elif preset_name == "Growth Accelerator":
        if col_name in ["revenue_cagr_5yr", "pat_cagr_5yr"]:
            if f_val is not None:
                apply_cell_style(cell, fill=GREEN_FILL if f_val >= 12.0 else RED_FILL, font=GREEN_FONT if f_val >= 12.0 else RED_FONT)
        elif col_name == "return_on_equity_pct":
            if f_val is not None:
                apply_cell_style(cell, fill=GREEN_FILL if f_val >= 12.0 else RED_FILL, font=GREEN_FONT if f_val >= 12.0 else RED_FONT)
        elif col_name == "operating_profit_margin_pct":
            if f_val is not None:
                apply_cell_style(cell, fill=GREEN_FILL if f_val >= 10.0 else RED_FILL, font=GREEN_FONT if f_val >= 10.0 else RED_FONT)
                
    # --- Preset 4: Dividend Champion ---
    elif preset_name == "Dividend Champion":
        if col_name == "dividend_yield_pct":
            if f_val is not None:
                apply_cell_style(cell, fill=GREEN_FILL if f_val >= 2.0 else RED_FILL, font=GREEN_FONT if f_val >= 2.0 else RED_FONT)
        elif col_name == "return_on_equity_pct":
            if f_val is not None:
                apply_cell_style(cell, fill=GREEN_FILL if f_val >= 10.0 else RED_FILL, font=GREEN_FONT if f_val >= 10.0 else RED_FONT)
        elif col_name == "debt_to_equity":
            if is_fin:
                apply_cell_style(cell, fill=GREEN_FILL, font=GREEN_FONT)
            elif f_val is not None:
                apply_cell_style(cell, fill=GREEN_FILL if f_val <= 1.2 else RED_FILL, font=GREEN_FONT if f_val <= 1.2 else RED_FONT)
                
    # --- Preset 5: Debt-Free Blue Chip ---
    elif preset_name == "Debt-Free Blue Chip":
        if col_name == "market_cap_crore":
            if f_val is not None:
                apply_cell_style(cell, fill=GREEN_FILL if f_val >= 20000.0 else RED_FILL, font=GREEN_FONT if f_val >= 20000.0 else RED_FONT)
        elif col_name == "return_on_equity_pct":
            if f_val is not None:
                apply_cell_style(cell, fill=GREEN_FILL if f_val >= 12.0 else RED_FILL, font=GREEN_FONT if f_val >= 12.0 else RED_FONT)
        elif col_name == "debt_to_equity":
            if is_df or (f_val is not None and f_val <= 0.05):
                apply_cell_style(cell, fill=GREEN_FILL, font=GREEN_FONT)
            else:
                apply_cell_style(cell, fill=RED_FILL, font=RED_FONT)
                
    # --- Preset 6: Turnaround Watch ---
    elif preset_name == "Turnaround Watch":
        if col_name == "debt_to_equity":
            # Check if D/E YoY actually declined
            declined = row_vals.get("de_declined_yoy", False)
            apply_cell_style(cell, fill=GREEN_FILL if declined else RED_FILL, font=GREEN_FONT if declined else RED_FONT)
        elif col_name == "return_on_equity_pct":
            if f_val is not None:
                apply_cell_style(cell, fill=GREEN_FILL if f_val >= 10.0 else RED_FILL, font=GREEN_FONT if f_val >= 10.0 else RED_FONT)
        elif col_name == "net_profit":
            if f_val is not None:
                apply_cell_style(cell, fill=GREEN_FILL if f_val >= 100.0 else RED_FILL, font=GREEN_FONT if f_val >= 100.0 else RED_FONT)


def generate_screener_excel(db_path: Union[str, Path], output_path: Union[str, Path], year: Union[str, int] = 2024):
    """
    Run all six stock screener presets, select the 20 KPI columns, sort them 
    by our composite score, and write them into separate sheets in the target 
    Excel file with professional styles and conditional green/red highlights.
    """
    out_file = Path(output_path)
    out_file.parent.mkdir(parents=True, exist_ok=True)
    
    # 1. Load the database and calculate composite scores for all companies
    df_scored = load_and_score_data(db_path, year=year)
    
    # 2. Run each preset filter on the scored data
    presets_dict = {
        "Quality Compounder": screen_quality_compounders(db_path, year=year),
        "Value Pick": screen_value_picks(db_path, year=year),
        "Growth Accelerator": screen_growth_accelerators(db_path, year=year),
        "Dividend Champion": screen_dividend_champions(db_path, year=year),
        "Debt-Free Blue Chip": screen_debt_free_blue_chips(db_path, year=year),
        "Turnaround Watch": screen_turnaround_watch(db_path, year=year)
    }
    
    # Create openpyxl workbook
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)
    
    for preset_name, df_filtered in presets_dict.items():
        print(f"Exporting {preset_name}: {len(df_filtered)} companies...")
        
        # Select 20 KPI columns (make sure they exist)
        existing_cols = [c for c in KPI_COLUMNS if c in df_filtered.columns]
        df_sheet = df_filtered[existing_cols].copy()
        
        # Sort sheet by composite quality score descending
        if "composite_quality_score" in df_sheet.columns:
            df_sheet.sort_values(by="composite_quality_score", ascending=False, inplace=True)
            
        # Add the turnaround flag to row values if on Turnaround Watch to help styling
        if preset_name == "Turnaround Watch" and "company_id" in df_sheet.columns:
            # Merge de_declined_yoy column from df_scored
            df_sheet = df_sheet.merge(
                df_scored[["company_id", "de_declined_yoy"]], 
                on="company_id", 
                how="left"
            )
            
        # Create a new sheet
        ws = wb.create_sheet(title=preset_name)
        
        # Write headers renamed for presentation
        headers = [get_column_rename_header(c) for c in existing_cols]
        ws.append(headers)
        
        # Write data rows
        for _, row in df_sheet.iterrows():
            row_data = []
            for col_name in existing_cols:
                row_data.append(row[col_name])
            ws.append(row_data)
            
        # Apply formatting and conditional formatting
        # Note: df_sheet contains the data we wrote (excluding de_declined_yoy from output rows)
        style_preset_sheet(ws, df_sheet[existing_cols], preset_name)
        
    # Save the workbook
    wb.save(out_file)
    print(f"\n[SUCCESS] Excel workbook successfully saved at: {out_file.resolve()}")
