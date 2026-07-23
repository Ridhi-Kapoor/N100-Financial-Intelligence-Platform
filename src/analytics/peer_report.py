"""
Peer Group Comparison Excel Report Generator Module.

This module generates a multi-worksheet Excel workbook at 'output/peer_comparison.xlsx'
containing 11 worksheets—one for each peer group.

Features per sheet:
- Columns: company_id, company_name, 20 financial KPI columns, and 20 percentile rank columns.
- Conditional formatting on percentile rank cells:
  - Green: >= 75th percentile
  - Yellow: 25th - 75th percentile
  - Red: < 25th percentile
- Benchmark company row highlighted with a gold/amber background.
- Summary row at the bottom showing the median value for each metric within the peer group.
- Styling powered by openpyxl.
"""

import logging
import sqlite3
from pathlib import Path
from typing import Dict, Optional, Tuple, Union

import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.analytics.ratios import calculate_roce

# Define project paths
PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
DEFAULT_DB_PATH = PROJECT_ROOT / "data" / "db" / "nifty100.db"
DEFAULT_OUTPUT_FILE = PROJECT_ROOT / "output" / "peer_comparison.xlsx"

# Configure logging
LOG_DIR = PROJECT_ROOT / "logs"
LOG_DIR.mkdir(parents=True, exist_ok=True)

logger = logging.getLogger("peer_report_analytics")
logger.setLevel(logging.INFO)

if not logger.handlers:
    file_handler = logging.FileHandler(
        LOG_DIR / "peer_report.log", mode="a", encoding="utf-8"
    )
    formatter = logging.Formatter(
        "%(asctime)s - %(name)s - %(levelname)s - %(message)s"
    )
    file_handler.setFormatter(formatter)
    logger.addHandler(file_handler)
    console_handler = logging.StreamHandler()
    console_handler.setFormatter(formatter)
    logger.addHandler(console_handler)

# 20 Selected KPI Columns & Presentation Labels
KPI_MAPPING: Dict[str, Tuple[str, bool]] = {
    # db_column_name: (presentation_label, higher_is_better)
    "sales": ("Revenue (Cr)", True),
    "net_profit": ("Net Profit (Cr)", True),
    "operating_profit": ("Operating Profit (Cr)", True),
    "market_cap_crore": ("Market Cap (Cr)", True),
    "net_profit_margin_pct": ("NPM %", True),
    "operating_profit_margin_pct": ("OPM %", True),
    "return_on_equity_pct": ("ROE %", True),
    "roce": ("ROCE %", True),
    "debt_to_equity": ("Debt-to-Equity", False),  # Lower is better
    "interest_coverage": ("Interest Coverage", True),
    "asset_turnover": ("Asset Turnover", True),
    "free_cash_flow_cr": ("FCF (Cr)", True),
    "cash_from_operations_cr": ("CFO (Cr)", True),
    "capex_cr": ("Capex (Cr)", True),
    "earnings_per_share": ("EPS (Rs)", True),
    "book_value_per_share": ("BVPS (Rs)", True),
    "revenue_cagr_5yr": ("Revenue CAGR 5Y %", True),
    "pat_cagr_5yr": ("PAT CAGR 5Y %", True),
    "eps_cagr_5yr": ("EPS CAGR 5Y %", True),
    "pe_ratio": ("P/E Ratio", False),  # Lower is better in value assessment
}

# Styling definitions
HEADER_FILL_PRIMARY = PatternFill(
    start_color="1F4E78", end_color="1F4E78", fill_type="solid"
)  # Dark Blue
HEADER_FILL_PERCENTILE = PatternFill(
    start_color="2E75B6", end_color="2E75B6", fill_type="solid"
)  # Steel Blue
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

BENCHMARK_FILL = PatternFill(
    start_color="FFE699", end_color="FFE699", fill_type="solid"
)  # Gold / Amber
BENCHMARK_FONT = Font(name="Calibri", size=11, bold=True, color="000000")

SUMMARY_FILL = PatternFill(
    start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
)  # Light Blue-Gray Accent
SUMMARY_FONT = Font(name="Calibri", size=11, bold=True, color="000000")

# Percentile Rank Conditional Formatting Fills & Fonts
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
GREEN_FONT = Font(name="Calibri", size=11, color="006100", bold=True)

YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
YELLOW_FONT = Font(name="Calibri", size=11, color="9C6500", bold=True)

RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
RED_FONT = Font(name="Calibri", size=11, color="9C0006", bold=True)

REGULAR_FONT = Font(name="Calibri", size=11)

THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

SUMMARY_BORDER = Border(
    left=Side(style="thin", color="808080"),
    right=Side(style="thin", color="808080"),
    top=Side(style="thin", color="808080"),
    bottom=Side(style="double", color="000000"),
)


def load_peer_comparison_dataset(
    db_path: Optional[Union[str, Path]] = None, year: Union[str, int] = "2024"
) -> pd.DataFrame:
    """
    Query database for peer groups and all 20 financial KPIs for target year.

    Args:
        db_path: Path to nifty100.db.
        year: Target year.

    Returns:
        pd.DataFrame containing company profiles, peer group data, and 20 raw KPI metrics.
    """
    path_db = Path(db_path) if db_path else DEFAULT_DB_PATH
    yr_str = str(year).strip()

    conn = sqlite3.connect(path_db)
    try:
        # Load peer groups
        pg = pd.read_sql(
            "SELECT company_id as id, peer_group_name, is_benchmark FROM peer_groups",
            conn,
        ).drop_duplicates(subset=["id"])
        pg["id"] = pg["id"].astype(str).str.strip()

        # Companies
        comps = pd.read_sql("SELECT id, company_name FROM companies", conn)
        comps["id"] = comps["id"].astype(str).str.strip()

        # Financial ratios
        fr = pd.read_sql(
            "SELECT company_id as id, net_profit_margin_pct, operating_profit_margin_pct, return_on_equity_pct, debt_to_equity, interest_coverage, asset_turnover, free_cash_flow_cr, capex_cr, earnings_per_share, book_value_per_share, cash_from_operations_cr, revenue_cagr_5yr, pat_cagr_5yr, eps_cagr_5yr, composite_quality_score FROM financial_ratios WHERE year = ?",
            conn,
            params=[yr_str],
        ).drop_duplicates(subset=["id"])
        if fr.empty:
            fr = pd.read_sql(
                "SELECT company_id as id, net_profit_margin_pct, operating_profit_margin_pct, return_on_equity_pct, debt_to_equity, interest_coverage, asset_turnover, free_cash_flow_cr, capex_cr, earnings_per_share, book_value_per_share, cash_from_operations_cr, revenue_cagr_5yr, pat_cagr_5yr, eps_cagr_5yr, composite_quality_score FROM financial_ratios",
                conn,
            ).drop_duplicates(subset=["id"], keep="last")
        fr["id"] = fr["id"].astype(str).str.strip()

        # Market Cap
        mc = pd.read_sql(
            "SELECT company_id as id, pe_ratio, pb_ratio, market_cap_crore FROM market_cap WHERE year = ?",
            conn,
            params=[int(yr_str) if yr_str.isdigit() else 2024],
        ).drop_duplicates(subset=["id"])
        if mc.empty:
            mc = pd.read_sql(
                "SELECT company_id as id, pe_ratio, pb_ratio, market_cap_crore FROM market_cap",
                conn,
            ).drop_duplicates(subset=["id"], keep="last")
        mc["id"] = mc["id"].astype(str).str.strip()

        # P&L
        pl = pd.read_sql(
            "SELECT company_id as id, sales, net_profit, operating_profit, profit_before_tax, interest FROM profitandloss WHERE year = ?",
            conn,
            params=[yr_str],
        ).drop_duplicates(subset=["id"])
        if pl.empty:
            pl = pd.read_sql(
                "SELECT company_id as id, sales, net_profit, operating_profit, profit_before_tax, interest FROM profitandloss",
                conn,
            ).drop_duplicates(subset=["id"], keep="last")
        pl["id"] = pl["id"].astype(str).str.strip()

        # Balance Sheet
        bs = pd.read_sql(
            "SELECT company_id as id, equity_capital, reserves, borrowings FROM balancesheet WHERE year = ?",
            conn,
            params=[yr_str],
        ).drop_duplicates(subset=["id"])
        if bs.empty:
            bs = pd.read_sql(
                "SELECT company_id as id, equity_capital, reserves, borrowings FROM balancesheet",
                conn,
            ).drop_duplicates(subset=["id"], keep="last")
        bs["id"] = bs["id"].astype(str).str.strip()

        # Sectors
        sec = pd.read_sql(
            "SELECT company_id as id, broad_sector FROM sectors", conn
        ).drop_duplicates(subset=["id"])
        sec["id"] = sec["id"].astype(str).str.strip()

        # Merge peer group records
        df = pd.merge(pg, comps, on="id", how="left")
        df = pd.merge(df, fr, on="id", how="left")
        df = pd.merge(df, mc, on="id", how="left")
        df = pd.merge(df, pl, on="id", how="left")
        df = pd.merge(df, bs, on="id", how="left")
        df = pd.merge(df, sec, on="id", how="left")

        # Compute ROCE
        roce_list = []
        for _, r in df.iterrows():
            try:
                pbt = (
                    float(r["profit_before_tax"])
                    if pd.notna(r.get("profit_before_tax"))
                    else 0.0
                )
                int_exp = float(r["interest"]) if pd.notna(r.get("interest")) else 0.0
                ebit = pbt + int_exp
                roce_res = calculate_roce(
                    ebit=ebit,
                    equity_capital=r.get("equity_capital"),
                    reserves=r.get("reserves"),
                    borrowings=r.get("borrowings"),
                    broad_sector=r.get("broad_sector"),
                )
                val = roce_res[0] if isinstance(roce_res, tuple) else roce_res
                roce_list.append(val)
            except Exception:
                roce_list.append(None)

        df["roce"] = roce_list
        return df
    finally:
        conn.close()


def generate_peer_comparison_excel(
    db_path: Optional[Union[str, Path]] = None,
    output_path: Optional[Union[str, Path]] = None,
    year: Union[str, int] = "2024",
) -> Path:
    """
    Generate output/peer_comparison.xlsx with 11 worksheets (one for each peer group).

    Args:
        db_path: Path to SQLite DB.
        output_path: Destination path for Excel file.
        year: Financial year.

    Returns:
        Path: Path to the generated Excel file.
    """
    out_file = Path(output_path) if output_path else DEFAULT_OUTPUT_FILE
    out_file.parent.mkdir(parents=True, exist_ok=True)

    logger.info("Loading dataset for peer comparison report...")
    df_all = load_peer_comparison_dataset(db_path=db_path, year=year)

    # Create OpenPyXL workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default initial sheet

    peer_groups = sorted([g for g in df_all["peer_group_name"].unique() if pd.notna(g)])
    logger.info(f"Generating Excel report for {len(peer_groups)} peer groups...")

    for pg_name in peer_groups:
        grp = df_all[df_all["peer_group_name"] == pg_name].copy()
        ws = wb.create_sheet(title=str(pg_name)[:31])  # Excel sheet title max 31 chars

        # 1. Calculate Percentile Ranks for all 20 KPI metrics within this peer group
        for col_name, (_, higher_is_better) in KPI_MAPPING.items():
            if col_name in grp.columns:
                series = grp[col_name].astype(float)
                grp[f"{col_name}_pct_rank"] = (
                    series.rank(
                        pct=True,
                        ascending=higher_is_better,
                        method="average",
                        na_option="keep",
                    )
                    * 100.0
                )

        # 2. Build Column Headers
        headers = ["company_id", "company_name"]
        header_labels = ["Company ID", "Company Name"]

        # KPI Value Headers
        for col_name in KPI_MAPPING:
            headers.append(col_name)
            header_labels.append(KPI_MAPPING[col_name][0])

        # KPI Percentile Rank Headers
        for col_name in KPI_MAPPING:
            headers.append(f"{col_name}_pct_rank")
            header_labels.append(f"{KPI_MAPPING[col_name][0]} (Pct Rank)")

        ws.append(header_labels)
        ws.row_dimensions[1].height = 28

        # Style Header Row
        for col_idx, col_key in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            is_pct_header = col_key.endswith("_pct_rank")
            fill_style = (
                HEADER_FILL_PERCENTILE if is_pct_header else HEADER_FILL_PRIMARY
            )
            cell.fill = fill_style
            cell.font = HEADER_FONT
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )

        # 3. Write Company Data Rows
        data_rows_start = 2
        for r_idx, (_, row) in enumerate(grp.iterrows(), start=data_rows_start):
            ws.row_dimensions[r_idx].height = 21
            is_benchmark = str(row.get("is_benchmark")).strip() == "True"

            row_data = [row.get("id"), row.get("company_name")]

            # Raw KPI Values
            for col_name in KPI_MAPPING:
                val = row.get(col_name)
                row_data.append(float(val) if (pd.notna(val) and val != "") else None)

            # Percentile Rank Values
            for col_name in KPI_MAPPING:
                rank_val = row.get(f"{col_name}_pct_rank")
                row_data.append(
                    float(rank_val) if (pd.notna(rank_val) and rank_val != "") else None
                )

            ws.append(row_data)

            # Style the data row cells
            for c_idx, col_key in enumerate(headers, start=1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.font = BENCHMARK_FONT if is_benchmark else REGULAR_FONT
                cell.border = THIN_BORDER

                # Alignment & Number Formatting
                if c_idx == 1:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif c_idx == 2:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                elif col_key.endswith("_pct_rank"):
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    cell.number_format = (
                        "0.0%"  # Format as percentage if 0-1, or 0.0 for rank
                    )
                    if cell.value is not None:
                        cell.number_format = "0.0"
                else:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    if col_key in [
                        "sales",
                        "net_profit",
                        "operating_profit",
                        "market_cap_crore",
                        "free_cash_flow_cr",
                        "cash_from_operations_cr",
                        "capex_cr",
                    ]:
                        cell.number_format = "#,##0.00"
                    else:
                        cell.number_format = "0.00"

                # Apply Row Background Fill & Conditional Formatting
                if is_benchmark:
                    cell.fill = BENCHMARK_FILL

                # Apply Conditional Formatting to Percentile Rank Cells
                if col_key.endswith("_pct_rank") and cell.value is not None:
                    rank = float(cell.value)
                    if rank >= 75.0:
                        cell.fill = GREEN_FILL
                        cell.font = GREEN_FONT
                    elif rank >= 25.0:
                        cell.fill = YELLOW_FILL
                        cell.font = YELLOW_FONT
                    else:
                        cell.fill = RED_FILL
                        cell.font = RED_FONT

        # 4. Add Summary Row at Bottom (Median for each metric)
        summary_row_idx = len(grp) + data_rows_start
        ws.row_dimensions[summary_row_idx].height = 24

        summary_data = ["PEER MEDIAN", "Peer Group Median"]

        # Calculate medians for KPI values
        for col_name in KPI_MAPPING:
            med = grp[col_name].dropna().median() if col_name in grp.columns else None
            summary_data.append(float(med) if (pd.notna(med)) else None)

        # Calculate medians for Percentile Ranks
        for col_name in KPI_MAPPING:
            rank_col = f"{col_name}_pct_rank"
            med_rank = (
                grp[rank_col].dropna().median() if rank_col in grp.columns else None
            )
            summary_data.append(float(med_rank) if (pd.notna(med_rank)) else None)

        ws.append(summary_data)

        # Style Summary Row
        for c_idx, col_key in enumerate(headers, start=1):
            cell = ws.cell(row=summary_row_idx, column=c_idx)
            cell.fill = SUMMARY_FILL
            cell.font = SUMMARY_FONT
            cell.border = SUMMARY_BORDER

            if c_idx == 1:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif c_idx == 2:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            elif col_key.endswith("_pct_rank"):
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = "0.0"
            else:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                if col_key in [
                    "sales",
                    "net_profit",
                    "operating_profit",
                    "market_cap_crore",
                    "free_cash_flow_cr",
                    "cash_from_operations_cr",
                    "capex_cr",
                ]:
                    cell.number_format = "#,##0.00"
                else:
                    cell.number_format = "0.00"

        # 5. Auto-fit column widths
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or "")
                if cell.row == 1:
                    max_len = max(max_len, len(val_str) // 2 + 6)
                else:
                    max_len = max(max_len, len(val_str) + 3)
            ws.column_dimensions[col_letter].width = max(max_len, 12)

    # Save workbook
    wb.save(out_file)
    logger.info(
        f"Peer comparison Excel report successfully created at: {out_file.resolve()}"
    )
    return out_file


def main() -> None:
    """Main execution block."""
    logger.info("Executing Peer Comparison Excel Generator...")
    out = generate_peer_comparison_excel()
    logger.info(f"Process complete! Output written to: {out}")


if __name__ == "__main__":
    main()
