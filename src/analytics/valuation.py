"""
Valuation Analytics Module (Day 26).

This module performs comprehensive valuation analysis across Nifty 100 companies:
- Calculates Free Cash Flow (FCF) Yield (%) using Free Cash Flow and Market Cap (in Crores).
- Computes sector median P/E ratios for each broad sector in the latest available financial year.
- Calculates 5-Year Median P/E per company across the latest 5 available financial years.
- Compares each company's P/E ratio against its sector median to determine:
    - PE_vs_sector_median_pct: Percentage difference relative to sector median.
    - Valuation Flag:
        * 'Caution': P/E > 1.5 * Sector Median
        * 'Discount': P/E < 0.7 * Sector Median
        * 'Fair': Otherwise
- Outputs:
    - output/valuation_summary.xlsx: Styled Excel workbook containing valuation metrics for all companies.
    - output/valuation_flags.csv: CSV file filtered for companies flagged as Caution or Discount.
"""

import logging
import math
import sqlite3
from pathlib import Path
from typing import Dict, Optional, Tuple, Union

import numpy as np
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import pandas as pd

from src.analytics.cashflow import calculate_free_cash_flow

# Define project paths
PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
DEFAULT_DB_PATH = PROJECT_ROOT / "data" / "db" / "nifty100.db"
DEFAULT_RAW_DIR = PROJECT_ROOT / "data" / "raw"
DEFAULT_PROCESSED_DIR = PROJECT_ROOT / "data" / "processed"
DEFAULT_OUTPUT_DIR = PROJECT_ROOT / "output"

DEFAULT_SUMMARY_EXCEL = DEFAULT_OUTPUT_DIR / "valuation_summary.xlsx"
DEFAULT_FLAGS_CSV = DEFAULT_OUTPUT_DIR / "valuation_flags.csv"

# Configure logging
LOG_DIR = PROJECT_ROOT / "logs"
LOG_DIR.mkdir(parents=True, exist_ok=True)

logger = logging.getLogger("valuation_analytics")
logger.setLevel(logging.INFO)

if not logger.handlers:
    file_handler = logging.FileHandler(
        LOG_DIR / "valuation_analytics.log", mode="a", encoding="utf-8"
    )
    formatter = logging.Formatter(
        "%(asctime)s - %(name)s - %(levelname)s - %(message)s"
    )
    file_handler.setFormatter(formatter)
    logger.addHandler(file_handler)
    console_handler = logging.StreamHandler()
    console_handler.setFormatter(formatter)
    logger.addHandler(console_handler)

# Openpyxl Styles
HEADER_FILL = PatternFill(
    start_color="1F4E78", end_color="1F4E78", fill_type="solid"
)  # Dark Blue
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

CAUTION_FILL = PatternFill(
    start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
)  # Soft Red
CAUTION_FONT = Font(name="Calibri", size=11, color="9C0006", bold=True)

DISCOUNT_FILL = PatternFill(
    start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
)  # Soft Green
DISCOUNT_FONT = Font(name="Calibri", size=11, color="006100", bold=True)

FAIR_FILL = PatternFill(
    start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
)  # Soft Yellow
FAIR_FONT = Font(name="Calibri", size=11, color="9C6500", bold=True)

REGULAR_FONT = Font(name="Calibri", size=11)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def calculate_fcf_yield(
    free_cash_flow: Optional[Union[float, int]],
    market_cap_crore: Optional[Union[float, int]],
) -> Optional[float]:
    """
    Calculate Free Cash Flow (FCF) Yield (%).

    Formula:
        FCF Yield (%) = (Free Cash Flow / Market Cap in Crores) * 100

    Args:
        free_cash_flow: Free Cash Flow in Crores.
        market_cap_crore: Market Capitalization in Crores.

    Returns:
        FCF Yield as a percentage float, or None if inputs are invalid or market cap <= 0.
    """
    if free_cash_flow is None or market_cap_crore is None:
        return None

    try:
        fcf_val = float(free_cash_flow)
        mcap_val = float(market_cap_crore)
    except (ValueError, TypeError) as e:
        logger.warning(
            f"FCF Yield calculation failed due to invalid type: fcf={free_cash_flow}, mcap={market_cap_crore}. Error: {e}"
        )
        return None

    if math.isnan(fcf_val) or math.isnan(mcap_val) or mcap_val <= 0:
        return None

    return (fcf_val / mcap_val) * 100.0


def calculate_sector_median_pe(
    df: pd.DataFrame, sector_col: str = "broad_sector", pe_col: str = "pe_ratio"
) -> Dict[str, float]:
    """
    Calculate the median P/E ratio for each broad sector.

    Args:
        df: DataFrame containing sector and P/E ratio columns.
        sector_col: Name of the sector column.
        pe_col: Name of the P/E ratio column.

    Returns:
        Dictionary mapping sector name to its median P/E ratio as a float.
    """
    if df.empty or sector_col not in df.columns or pe_col not in df.columns:
        logger.warning(
            f"Unable to calculate sector median P/E: required columns '{sector_col}' or '{pe_col}' missing."
        )
        return {}

    valid_df = df.copy()
    valid_df[pe_col] = pd.to_numeric(valid_df[pe_col], errors="coerce")
    valid_df = valid_df.dropna(subset=[sector_col, pe_col])

    medians = valid_df.groupby(sector_col)[pe_col].median().to_dict()
    return {str(k): float(v) for k, v in medians.items()}


def assign_valuation_flag(
    pe_ratio: Optional[Union[float, int]],
    sector_median_pe: Optional[Union[float, int]],
) -> str:
    """
    Compare a company's P/E ratio with its sector median and assign a valuation flag.

    Rules:
        - Caution: P/E > 1.5 * Sector Median
        - Discount: P/E < 0.7 * Sector Median
        - Fair: Otherwise (or if P/E or sector median is invalid/missing)

    Args:
        pe_ratio: Company P/E ratio.
        sector_median_pe: Sector median P/E ratio.

    Returns:
        Valuation Flag string: 'Caution', 'Discount', or 'Fair'.
    """
    if pe_ratio is None or sector_median_pe is None:
        return "Fair"

    try:
        pe_val = float(pe_ratio)
        med_val = float(sector_median_pe)
    except (ValueError, TypeError):
        return "Fair"

    if math.isnan(pe_val) or math.isnan(med_val) or med_val <= 0:
        return "Fair"

    if pe_val > 1.5 * med_val:
        return "Caution"
    elif pe_val < 0.7 * med_val:
        return "Discount"
    else:
        return "Fair"


def calculate_pe_vs_sector_median_pct(
    pe_ratio: Optional[Union[float, int]],
    sector_median_pe: Optional[Union[float, int]],
) -> Optional[float]:
    """
    Calculate percentage difference between company's P/E ratio and its sector median.

    Formula:
        PE_vs_sector_median_pct = ((Company P/E - Sector Median P/E) / Sector Median P/E) * 100

    Args:
        pe_ratio: Company P/E ratio.
        sector_median_pe: Sector median P/E ratio.

    Returns:
        Percentage difference float, or None if inputs are invalid or sector median <= 0.
    """
    if pe_ratio is None or sector_median_pe is None:
        return None

    try:
        pe_val = float(pe_ratio)
        med_val = float(sector_median_pe)
    except (ValueError, TypeError):
        return None

    if math.isnan(pe_val) or math.isnan(med_val) or med_val <= 0:
        return None

    return ((pe_val - med_val) / med_val) * 100.0


def calculate_5yr_median_pe(
    df_market_cap: pd.DataFrame,
    company_id: str,
    latest_year: Optional[int] = None,
    years_back: int = 5,
) -> Optional[float]:
    """
    Calculate the 5-Year median P/E ratio for a company using the latest available financial years.

    Args:
        df_market_cap: DataFrame containing 'company_id', 'year', and 'pe_ratio'.
        company_id: The target company ID / ticker.
        latest_year: Upper bound year (defaults to max year in dataset for that company).
        years_back: Number of years to include (default 5).

    Returns:
        Median P/E ratio over the 5-year window, or None if insufficient data.
    """
    if (
        df_market_cap.empty
        or "company_id" not in df_market_cap.columns
        or "pe_ratio" not in df_market_cap.columns
    ):
        return None

    comp_df = df_market_cap[
        df_market_cap["company_id"].astype(str).str.strip() == str(company_id).strip()
    ].copy()
    if comp_df.empty:
        return None

    comp_df["pe_ratio"] = pd.to_numeric(comp_df["pe_ratio"], errors="coerce")
    comp_df["year_num"] = pd.to_numeric(comp_df["year"], errors="coerce")
    comp_df = comp_df.dropna(subset=["pe_ratio", "year_num"])

    if comp_df.empty:
        return None

    if latest_year is None:
        latest_year = int(comp_df["year_num"].max())

    min_year = latest_year - (years_back - 1)
    window_df = comp_df[
        (comp_df["year_num"] >= min_year) & (comp_df["year_num"] <= latest_year)
    ]

    if window_df.empty:
        return None

    return float(window_df["pe_ratio"].median())


def load_valuation_datasets(
    db_path: Optional[Union[str, Path]] = None,
    raw_dir: Optional[Union[str, Path]] = None,
    processed_dir: Optional[Union[str, Path]] = None,
) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    Load financial datasets required for valuation analysis.

    Supports loading from SQLite database ('nifty100.db'), raw Excel ('market_cap.xlsx'),
    and processed CSV datasets with fallback mechanisms.

    Returns:
        Tuple of DataFrames: (market_cap, companies, sectors, cashflow, financial_ratios)
    """
    path_db = Path(db_path) if db_path else DEFAULT_DB_PATH
    path_raw = Path(raw_dir) if raw_dir else DEFAULT_RAW_DIR
    path_proc = Path(processed_dir) if processed_dir else DEFAULT_PROCESSED_DIR

    df_mc = pd.DataFrame()
    df_comp = pd.DataFrame()
    df_sec = pd.DataFrame()
    df_cf = pd.DataFrame()
    df_rat = pd.DataFrame()

    # 1. Try loading market_cap from raw Excel first as per prompt specification
    mc_excel = path_raw / "market_cap.xlsx"
    if mc_excel.exists():
        try:
            logger.info(f"Loading market_cap from Excel: {mc_excel}")
            df_mc = pd.read_excel(mc_excel)
        except Exception as e:
            logger.warning(f"Could not load market_cap from Excel '{mc_excel}': {e}")

    # 2. Try database loading if available
    if path_db.exists():
        try:
            logger.info(f"Connecting to SQLite database: {path_db}")
            conn = sqlite3.connect(path_db)
            if df_mc.empty:
                df_mc = pd.read_sql("SELECT * FROM market_cap", conn)
            df_comp = pd.read_sql("SELECT * FROM companies", conn)
            df_sec = pd.read_sql("SELECT * FROM sectors", conn)
            df_cf = pd.read_sql("SELECT * FROM cashflow", conn)
            df_rat = pd.read_sql("SELECT * FROM financial_ratios", conn)
            conn.close()
        except Exception as e:
            logger.warning(f"Error loading datasets from database '{path_db}': {e}")

    # 3. Fallback to processed CSV files for missing tables
    def read_csv_safe(file_path: Path) -> pd.DataFrame:
        """Safely load a processed CSV file skipping metadata headers if present."""
        if not file_path.exists():
            return pd.DataFrame()
        with open(file_path, "r", encoding="utf-8") as f:
            first_line = f.readline()
        if (
            first_line.startswith("id,")
            or first_line.startswith("company_id,")
            or "id," in first_line
        ):
            df = pd.read_csv(file_path)
        else:
            df = pd.read_csv(file_path, header=1)
        df.columns = [c.strip() for c in df.columns]
        return df

    if df_mc.empty:
        df_mc = read_csv_safe(path_proc / "market_cap.csv")
    if df_comp.empty:
        df_comp = read_csv_safe(path_proc / "companies.csv")
    if df_sec.empty:
        df_sec = read_csv_safe(path_proc / "sectors.csv")
    if df_cf.empty:
        df_cf = read_csv_safe(path_proc / "cashflow.csv")
    if df_rat.empty:
        df_rat = read_csv_safe(path_proc / "financial_ratios.csv")

    # Standardize company key names across DataFrames
    if (
        not df_comp.empty
        and "id" in df_comp.columns
        and "company_id" not in df_comp.columns
    ):
        df_comp.rename(columns={"id": "company_id"}, inplace=True)

    for df_item in [df_mc, df_comp, df_sec, df_cf, df_rat]:
        if not df_item.empty and "company_id" in df_item.columns:
            df_item["company_id"] = df_item["company_id"].astype(str).str.strip()
        if not df_item.empty and "year" in df_item.columns:
            df_item["year"] = pd.to_numeric(df_item["year"], errors="coerce")

    return df_mc, df_comp, df_sec, df_cf, df_rat


def run_valuation_analysis(
    db_path: Optional[Union[str, Path]] = None,
    raw_dir: Optional[Union[str, Path]] = None,
    processed_dir: Optional[Union[str, Path]] = None,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    Perform valuation analysis for all 92 companies.

    Steps:
    1. Load financial datasets (market_cap, companies, sectors, cashflow, financial_ratios).
    2. Identify the latest available financial year in market_cap (e.g., 2024).
    3. Calculate Free Cash Flow (FCF) for each company for the latest year using operating & investing cash flows.
    4. Calculate FCF Yield (%) = (Free Cash Flow / Market Cap in Crores) * 100.
    5. Calculate 5-Year Median P/E for each company using the latest 5 financial years.
    6. Compute the sector median P/E ratio for each broad sector in the latest year.
    7. Compare each company's P/E with its sector median to calculate:
        - PE vs Sector Median (%)
        - Valuation Flag ('Caution', 'Discount', 'Fair')
    8. Generate complete summary DataFrame and filtered flags DataFrame.

    Returns:
        Tuple of (df_summary, df_flags)
    """
    logger.info("Executing Valuation Analytics Pipeline...")
    df_mc, df_comp, df_sec, df_cf, df_rat = load_valuation_datasets(
        db_path, raw_dir, processed_dir
    )

    if df_mc.empty or df_sec.empty:
        raise ValueError(
            "Critical valuation datasets (market_cap or sectors) could not be loaded."
        )

    # 1. Determine latest available year in market_cap
    latest_year = int(df_mc["year"].dropna().max())
    logger.info(f"Latest available financial year in market_cap dataset: {latest_year}")

    # 2. Filter latest year market cap metrics
    df_mc_latest = (
        df_mc[df_mc["year"] == latest_year]
        .drop_duplicates(subset=["company_id"])
        .copy()
    )

    # 3. Clean company metadata & sector assignments
    df_sec_clean = df_sec.drop_duplicates(subset=["company_id"])[
        ["company_id", "broad_sector"]
    ].copy()

    if not df_comp.empty and "company_name" in df_comp.columns:
        df_comp_clean = df_comp.drop_duplicates(subset=["company_id"])[
            ["company_id", "company_name"]
        ].copy()
    else:
        df_comp_clean = pd.DataFrame(
            {
                "company_id": df_sec_clean["company_id"],
                "company_name": df_sec_clean["company_id"],
            }
        )

    # Merge base company records with sector and latest market_cap metrics
    base_df = pd.merge(df_sec_clean, df_comp_clean, on="company_id", how="inner")
    base_df = pd.merge(
        base_df,
        df_mc_latest[
            ["company_id", "market_cap_crore", "pe_ratio", "pb_ratio", "ev_ebitda"]
        ],
        on="company_id",
        how="left",
    )

    # 4. Calculate FCF for the latest year
    # Use calculate_free_cash_flow from cashflow dataset with fallback to financial_ratios.free_cash_flow_cr
    df_cf_latest = (
        df_cf[df_cf["year"] == latest_year]
        .drop_duplicates(subset=["company_id"])
        .copy()
        if not df_cf.empty
        else pd.DataFrame()
    )
    df_rat_latest = (
        df_rat[df_rat["year"] == latest_year]
        .drop_duplicates(subset=["company_id"])
        .copy()
        if not df_rat.empty
        else pd.DataFrame()
    )

    fcf_records = []
    for cid in base_df["company_id"]:
        fcf_val = None
        if not df_cf_latest.empty and cid in df_cf_latest["company_id"].values:
            row_cf = df_cf_latest[df_cf_latest["company_id"] == cid].iloc[0]
            op_act = row_cf.get("operating_activity")
            inv_act = row_cf.get("investing_activity")
            fcf_val = calculate_free_cash_flow(op_act, inv_act)

        if (
            fcf_val is None
            and not df_rat_latest.empty
            and cid in df_rat_latest["company_id"].values
        ):
            row_rat = df_rat_latest[df_rat_latest["company_id"] == cid].iloc[0]
            raw_fcf = row_rat.get("free_cash_flow_cr")
            if pd.notna(raw_fcf):
                try:
                    fcf_val = float(raw_fcf)
                except (ValueError, TypeError):
                    fcf_val = None

        fcf_records.append({"company_id": cid, "free_cash_flow": fcf_val})

    df_fcf = pd.DataFrame(fcf_records)
    base_df = pd.merge(base_df, df_fcf, on="company_id", how="left")

    # 5. Calculate FCF Yield (%)
    base_df["fcf_yield_pct"] = base_df.apply(
        lambda r: calculate_fcf_yield(r["free_cash_flow"], r["market_cap_crore"]),
        axis=1,
    )

    # 6. Calculate 5-Year Median P/E for each company
    pe_5yr_dict = {}
    for cid in base_df["company_id"]:
        pe_5yr_dict[cid] = calculate_5yr_median_pe(
            df_mc, cid, latest_year=latest_year, years_back=5
        )

    base_df["pe_5yr_median"] = base_df["company_id"].map(pe_5yr_dict)

    # 7. Calculate Sector Median P/E for the latest year
    sector_medians = calculate_sector_median_pe(
        base_df, sector_col="broad_sector", pe_col="pe_ratio"
    )
    logger.info(f"Sector Median P/E ratios ({latest_year}): {sector_medians}")

    base_df["sector_median_pe"] = base_df["broad_sector"].map(sector_medians)

    # 8. Calculate PE vs Sector Median (%) & Valuation Flags
    base_df["pe_vs_sector_median_pct"] = base_df.apply(
        lambda r: calculate_pe_vs_sector_median_pct(
            r["pe_ratio"], r["sector_median_pe"]
        ),
        axis=1,
    )
    base_df["valuation_flag"] = base_df.apply(
        lambda r: assign_valuation_flag(r["pe_ratio"], r["sector_median_pe"]), axis=1
    )

    # 9. Format output column headers and names exactly as requested
    summary_cols = [
        "Company ID",
        "Company Name",
        "Sector",
        "P/E",
        "P/B",
        "EV/EBITDA",
        "FCF Yield (%)",
        "5-Year Median P/E",
        "PE vs Sector Median (%)",
        "Valuation Flag",
    ]

    rename_map = {
        "company_id": "Company ID",
        "company_name": "Company Name",
        "broad_sector": "Sector",
        "pe_ratio": "P/E",
        "pb_ratio": "P/B",
        "ev_ebitda": "EV/EBITDA",
        "fcf_yield_pct": "FCF Yield (%)",
        "pe_5yr_median": "5-Year Median P/E",
        "pe_vs_sector_median_pct": "PE vs Sector Median (%)",
        "valuation_flag": "Valuation Flag",
    }

    df_summary = base_df.rename(columns=rename_map)[summary_cols].copy()

    # Round numerical metrics to 2 decimal places for clean presentation
    numeric_cols = [
        "P/E",
        "P/B",
        "EV/EBITDA",
        "FCF Yield (%)",
        "5-Year Median P/E",
        "PE vs Sector Median (%)",
    ]
    for col in numeric_cols:
        df_summary[col] = pd.to_numeric(df_summary[col], errors="coerce").round(2)

    # Filter for companies flagged as Caution or Discount
    df_flags = df_summary[
        df_summary["Valuation Flag"].isin(["Caution", "Discount"])
    ].copy()

    logger.info(
        f"Valuation analysis complete for {len(df_summary)} companies. "
        f"Flags assigned: {df_summary['Valuation Flag'].value_counts().to_dict()}"
    )

    return df_summary, df_flags


def export_valuation_results(
    df_summary: pd.DataFrame,
    df_flags: pd.DataFrame,
    output_dir: Optional[Union[str, Path]] = None,
) -> Tuple[Path, Path]:
    """
    Export valuation analysis results to Excel workbook and CSV file.

    Outputs:
        - output/valuation_summary.xlsx: Multi-styled Excel workbook.
        - output/valuation_flags.csv: CSV file with Caution and Discount companies.

    Args:
        df_summary: DataFrame with valuation metrics for all companies.
        df_flags: DataFrame with Caution and Discount companies.
        output_dir: Destination folder path (defaults to 'output/').

    Returns:
        Tuple of (excel_path, csv_path)
    """
    out_dir = Path(output_dir) if output_dir else DEFAULT_OUTPUT_DIR
    out_dir.mkdir(parents=True, exist_ok=True)

    excel_path = out_dir / "valuation_summary.xlsx"
    csv_path = out_dir / "valuation_flags.csv"

    logger.info(f"Exporting valuation summary Excel to: {excel_path}")

    # Create Openpyxl workbook for styled Excel export
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Valuation Summary"
    ws.views.sheetView[0].showGridLines = True

    # Write Header
    headers = list(df_summary.columns)
    ws.append(headers)

    for col_idx, header_text in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = THIN_BORDER

    ws.row_dimensions[1].height = 28

    # Write Rows with formatting
    for r_idx, row in df_summary.iterrows():
        row_values = list(row)
        ws.append(row_values)
        current_row_num = ws.max_row
        ws.row_dimensions[current_row_num].height = 20

        val_flag = str(row.get("Valuation Flag", "")).strip()

        for c_idx, val in enumerate(row_values, 1):
            cell = ws.cell(row=current_row_num, column=c_idx)
            cell.border = THIN_BORDER
            cell.font = REGULAR_FONT

            col_name = headers[c_idx - 1]

            # Text alignment
            if col_name in ["Company ID", "Company Name", "Sector"]:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            elif col_name == "Valuation Flag":
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if val_flag == "Caution":
                    cell.fill = CAUTION_FILL
                    cell.font = CAUTION_FONT
                elif val_flag == "Discount":
                    cell.fill = DISCOUNT_FILL
                    cell.font = DISCOUNT_FONT
                elif val_flag == "Fair":
                    cell.fill = FAIR_FILL
                    cell.font = FAIR_FONT
            else:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                # Number formatting
                if isinstance(val, (int, float, np.number)) and not math.isnan(val):
                    cell.number_format = "0.00"

    # Auto-adjust column widths
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            val_str = str(cell.value or "")
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    wb.save(excel_path)

    # Export CSV file
    logger.info(f"Exporting valuation flags CSV to: {csv_path}")
    df_flags.to_csv(csv_path, index=False, encoding="utf-8")

    logger.info("Valuation results successfully saved to disk.")
    return excel_path, csv_path


def main() -> None:
    """CLI Entry point for Day 26 Valuation Analytics Module."""
    print("=" * 60)
    print("      DAY 26: NIFTY 100 VALUATION ANALYTICS MODULE")
    print("=" * 60)

    try:
        df_summary, df_flags = run_valuation_analysis()
        excel_path, csv_path = export_valuation_results(df_summary, df_flags)

        print("\nPipeline Summary:")
        print(f"- Total Companies Evaluated: {len(df_summary)}")
        print(
            f"- Caution Flagged Companies: {(df_summary['Valuation Flag'] == 'Caution').sum()}"
        )
        print(
            f"- Discount Flagged Companies: {(df_summary['Valuation Flag'] == 'Discount').sum()}"
        )
        print(
            f"- Fair Flagged Companies: {(df_summary['Valuation Flag'] == 'Fair').sum()}"
        )
        print("\nGenerated Output Files:")
        print(f"  [1] Excel Summary: {excel_path.resolve()}")
        print(f"  [2] CSV Flags:     {csv_path.resolve()}")
        print("=" * 60)
    except Exception as e:
        logger.error(
            f"Error in Valuation Analytics Module execution: {e}", exc_info=True
        )
        print(f"Execution Error: {e}")


if __name__ == "__main__":
    main()
