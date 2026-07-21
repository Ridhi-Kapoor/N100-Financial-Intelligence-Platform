"""
Report Generation Script.

Triggers the Excel export module to generate `output/screener_output.xlsx`,
which contains worksheets for each preset with conditional green/red formatting
for active filters, and verifies its correctness.
"""

from pathlib import Path
import openpyxl
from src.screener.excel_export import generate_screener_excel

PROJECT_ROOT = Path(__file__).resolve().parent.parent
DB_FILE = PROJECT_ROOT / "data" / "db" / "nifty100.db"
OUTPUT_FILE = PROJECT_ROOT / "output" / "screener_output.xlsx"

def main():
    print("=" * 110)
    print("                       NIFTY 100 STOCK SCREENER - EXCEL REPORT GENERATION")
    print("=" * 110)
    print(f"Database File: {DB_FILE.resolve()}")
    print(f"Output File:   {OUTPUT_FILE.resolve()}")
    
    if not DB_FILE.exists():
        print(f"[ERROR] Database file not found at: {DB_FILE.resolve()}")
        return
        
    try:
        # 1. Generate the Excel file
        generate_screener_excel(DB_FILE, OUTPUT_FILE, year=2024)
        
        # 2. Verify and validate the generated Excel file
        print("\nVerifying generated Excel file...")
        if not OUTPUT_FILE.exists():
            print(f"[FAILED] Excel file was not created at: {OUTPUT_FILE.resolve()}")
            return
            
        wb = openpyxl.load_workbook(OUTPUT_FILE)
        sheet_names = wb.sheetnames
        print(f"Found {len(sheet_names)} worksheets in the file:")
        
        expected_sheets = [
            "Quality Compounder",
            "Value Pick",
            "Growth Accelerator",
            "Dividend Champion",
            "Debt-Free Blue Chip",
            "Turnaround Watch"
        ]
        
        all_ok = True
        for sheet in expected_sheets:
            if sheet in sheet_names:
                ws = wb[sheet]
                rows = ws.max_row
                cols = ws.max_column
                # Max row - 1 header = count of companies
                comp_count = rows - 1
                print(f"  - Sheet '{sheet}': {comp_count} companies, {cols} columns.")
                
                # Check that rows count is reasonable
                if comp_count < 5 or comp_count > 50:
                    print(f"    [WARNING] Company count {comp_count} is outside the expected 5 to 50 range.")
                    all_ok = False
            else:
                print(f"  - [MISSING] Sheet '{sheet}' is not in the workbook!")
                all_ok = False
                
        if all_ok:
            print("\n[SUCCESS] Excel workbook generated successfully with all preset sheets!")
            print(f"Verify formatting and layout inside: {OUTPUT_FILE.resolve()}")
        else:
            print("\n[WARNING] Excel workbook generated, but some sheets failed validation checks.")
            
    except Exception as e:
        print(f"[ERROR] Failed to generate Excel report: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()
